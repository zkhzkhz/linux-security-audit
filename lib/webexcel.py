#!/usr/bin/env python3
"""
Web Excel Viewer with Basic Authentication.
Serves Excel audit reports via web with password protection.
"""
import os
import sys
import argparse
from pathlib import Path
from functools import wraps
import base64

from flask import Flask, request, Response, send_file, render_template_string, redirect, url_for
import openpyxl
from openpyxl.utils import get_column_letter

app = Flask(__name__)

# Configuration
AUTH_USERNAME = os.environ.get("WEBEXCEL_USER", "admin")
AUTH_PASSWORD = os.environ.get("WEBEXCEL_PASS", "audit2024")
REPORTS_BASE = Path("/root/linux-security-audit/reports")

# HTML Templates
INDEX_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>Security Audit Reports</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 20px; background: #f5f5f5; }
        h1 { color: #333; }
        .report-card {
            background: white;
            border: 1px solid #ddd;
            border-radius: 8px;
            padding: 15px;
            margin: 10px 0;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        .report-card h3 { margin: 0 0 10px 0; color: #366092; }
        .report-card p { margin: 5px 0; color: #666; }
        .btn {
            display: inline-block;
            padding: 8px 16px;
            background: #366092;
            color: white;
            text-decoration: none;
            border-radius: 4px;
            margin-right: 10px;
        }
        .btn:hover { background: #2a4a72; }
        .stats { font-size: 12px; color: #888; }
    </style>
</head>
<body>
    <h1>Security Audit Reports</h1>
    <p><a href="/logout" class="btn">Logout</a></p>

    {% for report in reports %}
    <div class="report-card">
        <h3>{{ report.name }}</h3>
        <p>{{ report.host }} | {{ report.arch }} | {{ report.generated }}</p>
        <p class="stats">{{ report.summary }}</p>
        <p>
            <a href="/view/{{ report.id }}" class="btn">View Online</a>
            <a href="/download/{{ report.id }}" class="btn">Download XLSX</a>
        </p>
    </div>
    {% endfor %}
</body>
</html>
"""

VIEW_TEMPLATE = """
<!DOCTYPE html>
<html>
<head>
    <title>{{ title }} - Audit Report</title>
    <style>
        body { font-family: Arial, sans-serif; margin: 0; background: #f5f5f5; }
        .header {
            background: #366092;
            color: white;
            padding: 15px 20px;
        }
        .header h1 { margin: 0; }
        .header a { color: white; margin-right: 15px; }
        .tabs {
            background: white;
            padding: 10px 20px;
            border-bottom: 1px solid #ddd;
        }
        .tabs a {
            display: inline-block;
            padding: 8px 16px;
            margin-right: 5px;
            text-decoration: none;
            border-radius: 4px 4px 0 0;
            background: #eee;
            color: #333;
        }
        .tabs a.active {
            background: #366092;
            color: white;
        }
        .content {
            padding: 20px;
            background: white;
            margin: 0 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        table {
            border-collapse: collapse;
            width: 100%;
            font-size: 12px;
        }
        th, td {
            border: 1px solid #ddd;
            padding: 6px 10px;
            text-align: left;
        }
        th {
            background: #366092;
            color: white;
            position: sticky;
            top: 0;
        }
        tr:nth-child(even) { background: #f9f9f9; }
        tr:hover { background: #f0f0f0; }
        .critical { background: #ffcccc !important; }
        .high { background: #ffe6cc !important; }
        .medium { background: #ffffcc !important; }
        .low { background: #ccffcc !important; }
        .fp { background: #e0e0e0 !important; }
        .secret { font-family: monospace; font-size: 11px; word-break: break-all; }
        .pagination {
            text-align: center;
            padding: 20px;
        }
        .pagination a {
            display: inline-block;
            padding: 8px 16px;
            margin: 0 5px;
            background: #366092;
            color: white;
            text-decoration: none;
            border-radius: 4px;
        }
        .pagination span {
            display: inline-block;
            padding: 8px 16px;
            margin: 0 5px;
            background: #eee;
            border-radius: 4px;
        }
    </style>
</head>
<body>
    <div class="header">
        <h1>{{ title }}</h1>
        <a href="/">Home</a>
        <a href="/download/{{ report_id }}">Download XLSX</a>
    </div>

    <div class="tabs">
        {% for sheet in sheets %}
        <a href="/view/{{ report_id }}/{{ sheet }}?page={{ page }}{% if filter_fp %}&fp=1{% endif %}"
           class="{% if sheet == current_sheet %}active{% endif %}">{{ sheet }}</a>
        {% endfor %}
    </div>

    <div class="content">
        <p>
            <label><input type="checkbox" onclick="toggleFP()" {% if filter_fp %}checked{% endif %}> Hide False Positives</label>
            <span style="float:right">Total: {{ total_rows }} rows | Showing: {{ rows|length }} rows</span>
        </p>

        <table>
            <tr>
                {% for header in headers %}
                <th>{{ header }}</th>
                {% endfor %}
            </tr>
            {% for row in rows %}
            <tr class="{{ row.css_class }}">
                {% for cell in row.cells %}
                <td class="{% if cell.is_secret %}secret{% endif %}">{{ cell.value }}</td>
                {% endfor %}
            </tr>
            {% endfor %}
        </table>

        <div class="pagination">
            {% if page > 1 %}
            <a href="/view/{{ report_id }}/{{ current_sheet }}?page={{ page - 1 }}{% if filter_fp %}&fp=1{% endif %}">Previous</a>
            {% endif %}
            <span>Page {{ page }} of {{ total_pages }}</span>
            {% if page < total_pages %}
            <a href="/view/{{ report_id }}/{{ current_sheet }}?page={{ page + 1 }}{% if filter_fp %}&fp=1{% endif %}">Next</a>
            {% endif %}
        </div>
    </div>

    <script>
    function toggleFP() {
        var url = new URL(window.location.href);
        if (url.searchParams.has('fp')) {
            url.searchParams.delete('fp');
        } else {
            url.searchParams.set('fp', '1');
        }
        window.location.href = url.toString();
    }
    </script>
</body>
</html>
"""

def check_auth(username, password):
    """Check if a username/password combination is valid."""
    return username == AUTH_USERNAME and password == AUTH_PASSWORD

def authenticate():
    """Send a 401 response that enables basic auth."""
    return Response(
        'Authentication required.\n'
        'Username: {}\n'.format(AUTH_USERNAME),
        401,
        {'WWW-Authenticate': 'Basic realm="Audit Reports"'}
    )

def requires_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.authorization
        if not auth or not check_auth(auth.username, auth.password):
            return authenticate()
        return f(*args, **kwargs)
    return decorated

def find_reports():
    """Find all available audit reports."""
    reports = []

    for report_dir in REPORTS_BASE.iterdir():
        if not report_dir.is_dir():
            continue

        # Method 1: Look for xlsx directly in report_dir/reports/
        direct_xlsx = report_dir / "reports" / "detailed-audit-findings.xlsx"
        if direct_xlsx.exists():
            summary_file = report_dir / "reports" / "summary.json"
            report = {
                "id": str(direct_xlsx.relative_to(REPORTS_BASE)),
                "path": str(direct_xlsx),
                "name": report_dir.name,
                "host": "unknown",
                "arch": "unknown",
                "generated": "unknown",
                "summary": ""
            }
            if summary_file.exists():
                try:
                    import json
                    with open(summary_file) as f:
                        data = json.load(f)
                        report["host"] = data.get("host", "unknown")
                        report["arch"] = data.get("arch", "unknown")
                        report["generated"] = data.get("generated", "unknown")[:19] if data.get("generated") else "unknown"
                except:
                    pass
            reports.append(report)
            continue

        # Method 2: Look for xlsx in scan subdirectories
        for scan_dir in report_dir.rglob("*"):
            if not scan_dir.is_dir():
                continue

            excel_file = scan_dir / "detailed-audit-findings.xlsx"
            if excel_file.exists():
                summary_file = scan_dir / "summary.json"
                report = {
                    "id": str(excel_file.relative_to(REPORTS_BASE)),
                    "path": str(excel_file),
                    "name": report_dir.name,
                    "host": "unknown",
                    "arch": "unknown",
                    "generated": "unknown",
                    "summary": ""
                }

                if summary_file.exists():
                    try:
                        import json
                        with open(summary_file) as f:
                            data = json.load(f)
                            report["host"] = data.get("host", "unknown")
                            report["arch"] = data.get("arch", "unknown")
                            report["generated"] = data.get("generated", "unknown")[:19] if data.get("generated") else "unknown"

                            # Build summary
                            modules = data.get("modules", [])
                            parts = []
                            for m in modules:
                                name = m.get("module", "unknown")
                                status = m.get("status", "unknown")
                                counts = m.get("counts", {})
                                total = sum(v for v in counts.values() if isinstance(v, int))
                                parts.append(f"{name}: {total}")
                            report["summary"] = " | ".join(parts)
                    except:
                        pass

                reports.append(report)

    return sorted(reports, key=lambda x: x["generated"], reverse=True)

@app.route("/")
@requires_auth
def index():
    reports = find_reports()
    return render_template_string(INDEX_TEMPLATE, reports=reports)

@app.route("/logout")
def logout():
    """Logout by forcing auth failure."""
    return authenticate()

@app.route("/download/<path:report_id>")
@requires_auth
def download(report_id):
    report_path = REPORTS_BASE / report_id
    if not report_path.exists():
        return "Report not found", 404

    return send_file(
        report_path,
        as_attachment=True,
        download_name=report_path.name
    )

@app.route("/view/<path:report_id>")
@requires_auth
def view_report(report_id):
    # Use first sheet name from the file
    report_path = REPORTS_BASE / report_id
    if report_path.exists():
        try:
            wb = openpyxl.load_workbook(report_path, read_only=True)
            first_sheet = wb.sheetnames[0] if wb.sheetnames else "Sheet1"
            wb.close()
            return redirect(url_for("view_sheet", report_id=report_id, sheet_name=first_sheet))
        except:
            pass
    return redirect(url_for("view_sheet", report_id=report_id, sheet_name="Sheet1"))

@app.route("/view/<path:report_id>/<sheet_name>")
@requires_auth
def view_sheet(report_id, sheet_name):
    report_path = REPORTS_BASE / report_id

    # Debug logging
    print(f"DEBUG: report_id={report_id}")
    print(f"DEBUG: report_path={report_path}")
    print(f"DEBUG: exists={report_path.exists()}")
    print(f"DEBUG: is_file={report_path.is_file() if report_path.exists() else 'N/A'}")

    if not report_path.exists():
        return f"Report not found: {report_path}", 404

    # Ensure it's a file, not directory
    if not report_path.is_file():
        return f"Not a file: {report_path}", 400

    page = request.args.get("page", 1, type=int)
    filter_fp = request.args.get("fp", "0") == "1"
    per_page = 100

    try:
        wb = openpyxl.load_workbook(str(report_path), read_only=True, data_only=True)

        # Get sheet names
        sheets = wb.sheetnames

        # Find the sheet
        ws = None
        for s in sheets:
            if s == sheet_name:
                ws = wb[s]
                break

        if ws is None:
            ws = wb[sheets[0]]
            sheet_name = sheets[0]

        # Read headers
        headers = []
        for cell in next(ws.iter_rows(min_row=1, max_row=1)):
            headers.append(str(cell.value) if cell.value else "")

        # Read data
        all_rows = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            cells = []
            is_fp = False
            severity = ""

            for col_idx, cell in enumerate(row):
                value = str(cell.value) if cell.value is not None else ""
                cells.append({
                    "value": value[:500] if len(value) > 500 else value,
                    "is_secret": col_idx in [6, 7]  # Full secret column
                })

                # Check severity and FP
                if col_idx == 1:  # Severity column
                    severity = value.lower()
                if col_idx == 8:  # FP reason column
                    if value:
                        is_fp = True

            # Determine CSS class
            css_class = ""
            if is_fp:
                css_class = "fp"
            elif severity == "critical":
                css_class = "critical"
            elif severity == "high":
                css_class = "high"
            elif severity == "medium":
                css_class = "medium"
            elif severity == "low":
                css_class = "low"

            all_rows.append({
                "cells": cells,
                "css_class": css_class,
                "is_fp": is_fp
            })

        wb.close()

        # Filter FPs if requested
        if filter_fp:
            all_rows = [r for r in all_rows if not r["is_fp"]]

        # Pagination
        total_rows = len(all_rows)
        total_pages = max(1, (total_rows + per_page - 1) // per_page)
        page = max(1, min(page, total_pages))
        start = (page - 1) * per_page
        rows = all_rows[start:start + per_page]

        # Get report title
        title = report_path.parent.parent.name

        return render_template_string(
            VIEW_TEMPLATE,
            title=title,
            report_id=report_id,
            sheets=sheets,
            current_sheet=sheet_name,
            headers=headers,
            rows=rows,
            page=page,
            total_pages=total_pages,
            total_rows=total_rows,
            filter_fp=filter_fp
        )

    except Exception as e:
        return f"Error loading report: {e}", 500

def main():
    parser = argparse.ArgumentParser(description="Web Excel Viewer for Audit Reports")
    parser.add_argument("--port", "-p", type=int, default=8080, help="Port to listen on")
    parser.add_argument("--host", "-H", default="0.0.0.0", help="Host to bind to")
    parser.add_argument("--user", "-u", default="admin", help="Username for authentication")
    parser.add_argument("--password", "-P", default=None, help="Password for authentication")
    args = parser.parse_args()

    global AUTH_USERNAME, AUTH_PASSWORD
    AUTH_USERNAME = args.user
    if args.password:
        AUTH_PASSWORD = args.password

    print(f"""
╔════════════════════════════════════════════════════════════╗
║           Security Audit Report Web Viewer                  ║
╠════════════════════════════════════════════════════════════╣
║  URL:      http://{args.host}:{args.port}/                              ║
║  Username: {AUTH_USERNAME:<45}║
║  Password: {AUTH_PASSWORD:<45}║
╠════════════════════════════════════════════════════════════╣
║  WARNING: Contains sensitive data (secrets, credentials)    ║
║  Do not expose to untrusted networks                        ║
╚════════════════════════════════════════════════════════════╝
    """)

    app.run(host=args.host, port=args.port, debug=False)

if __name__ == "__main__":
    main()
