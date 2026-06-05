#!/usr/bin/env python3
"""
Generate detailed Excel report with actual secrets for verification.
"""
import json
import sys
from pathlib import Path
from datetime import datetime
from collections import defaultdict
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Default scan directory (can be overridden via command line or by detecting latest)
DEFAULT_SCAN_DIR = Path("/root/linux-security-audit/reports/liteserver-mindie-audit-original/reports/xihe-910b2-prod-node-6-20260604-061752")
DEFAULT_OUTPUT_DIR = Path("/root/linux-security-audit/reports/liteserver-mindie-audit-original")

# Styles
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CRITICAL_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
HIGH_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
MEDIUM_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LOW_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
FP_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def sanitize_string(s, max_len=5000):
    """Remove illegal characters for Excel and truncate."""
    if not s:
        return ""
    s = str(s)[:max_len]
    # Remove control characters and other illegal chars for Excel
    illegal_chars = []
    for i in range(0, 32):
        if i not in (9, 10, 13):  # Allow tab, newline, carriage return
            illegal_chars.append(chr(i))
    for char in illegal_chars:
        s = s.replace(char, '')
    return s

def load_json_lines(path):
    """Load JSONL file (one JSON per line)."""
    results = []
    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        for line in f:
            line = line.strip()
            if line and line != '[]':
                try:
                    results.append(json.loads(line))
                except:
                    pass
    return results

def load_json(path):
    with open(path, 'r', encoding='utf-8', errors='replace') as f:
        return json.load(f)

def is_false_positive(finding, source="gitleaks"):
    """Determine if a finding is likely a false positive."""
    # Get file path - handle both gitleaks and trufflehog formats
    where = finding.get("File", "") or finding.get("where", "")
    if not where:
        where = finding.get("SourceMetadata", {}).get("Data", {}).get("Filesystem", {}).get("file", "")

    note = finding.get("note", "") or finding.get("Description", "")
    title = finding.get("title", "") or finding.get("DetectorName", "") or finding.get("RuleID", "")
    severity = finding.get("severity", "")

    # FP patterns - ordered by specificity
    # 1. Docker镜像层系统文件 (最大误报来源)
    if "/overlay2/" in where:
        if any(x in where for x in ["/usr/share/doc/", "/usr/share/misc/", "/usr/lib/", "/var/cache/dnf/"]):
            return "Docker镜像层系统文件"
        if "/usr/local/go/" in where and ("testdata" in where or "test" in where):
            return "Docker镜像层Go测试数据"
        if "/home/mindspore/miniconda/" in where and "/lib/python" in where:
            return "Docker镜像层Python stdlib"
        if any(x in where for x in ["/usr/include/", "/usr/local/lib/"]):
            return "Docker镜像层编译产物"

    # 2. Python标准库源码误匹配 (gitcode-token-rule, mysql-cnf-password等)
    if "/lib/python" in where or "/lib64/python" in where:
        if any(x in where for x in ["urllib/request.py", "distutils/", "nntplib.py", "hashlib.py"]):
            return "Python stdlib源码误匹配"
        if "/test/" in where:
            return "Python stdlib测试代码"

    # 3. 二进制文件误报
    if any(where.endswith(ext) for ext in [".mgc", ".solv", ".db", ".bolt", ".bin"]):
        return "二进制文件误匹配"

    # 4. 测试/示例数据
    if "/test/" in where or "/tests/" in where or "/testdata/" in where:
        return "测试数据"
    if "/examplefiles/" in where or "/examples/" in where:
        return "示例文件"
    if "_test.py" in where or "test_" in where:
        return "测试代码"
    if "fixture" in where.lower():
        return "测试fixture"

    # 5. 文档文件
    if "/usr/share/doc/" in where or "/docs/" in where:
        return "文档文件"

    # 6. Go stdlib测试数据
    if "/usr/local/go/" in where:
        return "Go stdlib文件"

    # 7. 规则特定误报
    if title == "gitcode-token-rule" and "urllib/request.py" in where:
        return "gitcode-token误匹配Python stdlib"
    if title == "mysql-cnf-password" and "/lib/python" in where:
        return "mysql-cnf误匹配Python stdlib"

    # 8. 已标记为FP
    if finding.get("is_likely_fp", False):
        return "自动识别误报"

    # 9. Check entropy for gitleaks format
    entropy = finding.get("Entropy", 0)
    if entropy and entropy < 3.0:
        return "低熵值"

    # 10. Check note for entropy
    if "entropy=" in note:
        try:
            entropy_val = float(note.split("entropy=")[1].split()[0])
            if entropy_val < 3.0:
                return "低熵值"
        except:
            pass

    return None

def create_excel_report(scan_dir=None, output_dir=None):
    """Create detailed Excel report with all findings."""
    # Use provided paths or defaults
    global SCAN_DIR, OUTPUT_DIR
    if scan_dir:
        SCAN_DIR = Path(scan_dir)
    if output_dir:
        OUTPUT_DIR = Path(output_dir)

    wb = openpyxl.Workbook()

    # Sheet 1: Gitleaks findings
    ws1 = wb.active
    ws1.title = "Gitleaks敏感信息"

    # Load gitleaks raw data
    gitleaks_raw_path = SCAN_DIR / "sensitive-info-scan/raw.json"
    gitleaks_findings = []

    if gitleaks_raw_path.exists():
        data = load_json(gitleaks_raw_path)
        if isinstance(data, list):
            gitleaks_findings = data

    # Headers
    headers = ["序号", "严重度", "类型", "文件路径", "行号", "密钥前缀", "完整密钥", "评分信息", "误报原因", "确认状态"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    # Data rows
    row = 2
    for i, f in enumerate(gitleaks_findings[:5000], 1):  # Limit to 5000
        # Gitleaks format: File, StartLine, Secret, RuleID, Entropy, Match, Description
        # Also handle result.json format: where, severity, title, note, secret_prefix
        file_path = f.get("File", "") or f.get("where", "")
        line_num = f.get("StartLine", "") or ""
        if not line_num and ":" in file_path:
            parts = file_path.rsplit(":", 1)
            file_path = parts[0]
            line_num = parts[1]

        secret = f.get("Secret", "") or f.get("secret_prefix", "") or f.get("Match", "")
        title = f.get("RuleID", "") or f.get("title", "")
        severity = f.get("severity", "")
        if not severity:
            # Gitleaks doesn't have severity, calculate based on entropy
            entropy = f.get("Entropy", 0)
            if entropy >= 4.5:
                severity = "high"
            elif entropy >= 3.5:
                severity = "medium"
            else:
                severity = "low"

        note = f.get("note", "") or f.get("Description", "")
        entropy_str = f"entropy={f.get('Entropy', 0):.2f}" if f.get("Entropy") else ""
        fp_reason = is_false_positive(f, "gitleaks")

        ws1.cell(row=row, column=1, value=i).border = THIN_BORDER
        ws1.cell(row=row, column=2, value=severity).border = THIN_BORDER
        ws1.cell(row=row, column=3, value=sanitize_string(title)).border = THIN_BORDER
        ws1.cell(row=row, column=4, value=sanitize_string(file_path)).border = THIN_BORDER
        ws1.cell(row=row, column=5, value=line_num).border = THIN_BORDER
        ws1.cell(row=row, column=6, value=sanitize_string(secret[:50] if secret else "")).border = THIN_BORDER
        ws1.cell(row=row, column=7, value=sanitize_string(secret)).border = THIN_BORDER
        ws1.cell(row=row, column=8, value=sanitize_string(f"{note} {entropy_str}".strip()[:100])).border = THIN_BORDER
        ws1.cell(row=row, column=9, value=sanitize_string(fp_reason or "")).border = THIN_BORDER
        ws1.cell(row=row, column=10, value="待确认").border = THIN_BORDER

        # Color by severity
        if fp_reason:
            for col in range(1, 11):
                ws1.cell(row=row, column=col).fill = FP_FILL
        elif severity == "critical":
            for col in range(1, 11):
                ws1.cell(row=row, column=col).fill = CRITICAL_FILL
        elif severity == "high":
            for col in range(1, 11):
                ws1.cell(row=row, column=col).fill = HIGH_FILL
        elif severity == "medium":
            for col in range(1, 11):
                ws1.cell(row=row, column=col).fill = MEDIUM_FILL

        row += 1

    # Adjust column widths
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 25
    ws1.column_dimensions['D'].width = 60
    ws1.column_dimensions['E'].width = 8
    ws1.column_dimensions['F'].width = 30
    ws1.column_dimensions['G'].width = 80
    ws1.column_dimensions['H'].width = 40
    ws1.column_dimensions['I'].width = 20
    ws1.column_dimensions['J'].width = 12

    # Sheet 2: TruffleHog findings
    ws2 = wb.create_sheet("TruffleHog密钥")

    trufflehog_path = SCAN_DIR / "sensitive-info-scan/trufflehog/filesystem.json"
    trufflehog_findings = []

    if trufflehog_path.exists():
        trufflehog_findings = load_json_lines(trufflehog_path)

    # Headers
    headers2 = ["序号", "检测器", "文件路径", "行号", "密钥前缀", "完整密钥", "脱敏密钥", "验证状态", "误报原因", "确认状态"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    row = 2
    for i, f in enumerate(trufflehog_findings, 1):
        detector = f.get("DetectorName", "")
        source_meta = f.get("SourceMetadata", {}).get("Data", {}).get("Filesystem", {})
        file_path = source_meta.get("file", "")
        line_num = source_meta.get("line", "")
        raw = f.get("Raw", "")
        redacted = f.get("Redacted", "")
        verified = f.get("Verified", False)
        fp_reason = is_false_positive(f, "trufflehog")

        ws2.cell(row=row, column=1, value=i).border = THIN_BORDER
        ws2.cell(row=row, column=2, value=sanitize_string(detector)).border = THIN_BORDER
        ws2.cell(row=row, column=3, value=sanitize_string(file_path)).border = THIN_BORDER
        ws2.cell(row=row, column=4, value=line_num).border = THIN_BORDER
        ws2.cell(row=row, column=5, value=sanitize_string(raw[:50] if raw else "")).border = THIN_BORDER
        ws2.cell(row=row, column=6, value=sanitize_string(raw)).border = THIN_BORDER
        ws2.cell(row=row, column=7, value=sanitize_string(redacted)).border = THIN_BORDER
        ws2.cell(row=row, column=8, value="已验证" if verified else "未验证").border = THIN_BORDER
        ws2.cell(row=row, column=9, value=sanitize_string(fp_reason or "")).border = THIN_BORDER
        ws2.cell(row=row, column=10, value="待确认").border = THIN_BORDER

        if fp_reason:
            for col in range(1, 11):
                ws2.cell(row=row, column=col).fill = FP_FILL

        row += 1

    # Adjust widths
    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 60
    ws2.column_dimensions['D'].width = 8
    ws2.column_dimensions['E'].width = 30
    ws2.column_dimensions['F'].width = 100
    ws2.column_dimensions['G'].width = 50
    ws2.column_dimensions['H'].width = 12
    ws2.column_dimensions['I'].width = 20
    ws2.column_dimensions['J'].width = 12

    # Sheet 3: Privilege Escalation
    ws3 = wb.create_sheet("提权逃逸检查")

    privesc_path = SCAN_DIR / "privesc-escape-check/result.json"
    privesc_findings = []

    if privesc_path.exists():
        data = load_json(privesc_path)
        privesc_findings = data.get("findings", [])

    headers3 = ["序号", "严重度", "类型", "位置", "说明", "确认状态"]
    for col, header in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    row = 2
    for i, f in enumerate(privesc_findings, 1):
        severity = f.get("severity", "")
        ws3.cell(row=row, column=1, value=i).border = THIN_BORDER
        ws3.cell(row=row, column=2, value=severity).border = THIN_BORDER
        ws3.cell(row=row, column=3, value=sanitize_string(f.get("title", ""))).border = THIN_BORDER
        ws3.cell(row=row, column=4, value=sanitize_string(f.get("where", ""))).border = THIN_BORDER
        ws3.cell(row=row, column=5, value=sanitize_string(f.get("note", "")[:100])).border = THIN_BORDER
        ws3.cell(row=row, column=6, value="待确认").border = THIN_BORDER

        if severity == "critical":
            for col in range(1, 7):
                ws3.cell(row=row, column=col).fill = CRITICAL_FILL
        elif severity == "high":
            for col in range(1, 7):
                ws3.cell(row=row, column=col).fill = HIGH_FILL

        row += 1

    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 12
    ws3.column_dimensions['C'].width = 30
    ws3.column_dimensions['D'].width = 80
    ws3.column_dimensions['E'].width = 60
    ws3.column_dimensions['F'].width = 12

    # Sheet 4: Lateral Movement
    ws4 = wb.create_sheet("横向移动风险")

    lateral_path = SCAN_DIR / "lateral-movement-scan/result.json"
    lateral_findings = []

    if lateral_path.exists():
        data = load_json(lateral_path)
        lateral_findings = data.get("findings", [])

    headers4 = ["序号", "严重度", "类型", "位置", "说明", "确认状态"]
    for col, header in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    row = 2
    for i, f in enumerate(lateral_findings, 1):
        ws4.cell(row=row, column=1, value=i).border = THIN_BORDER
        ws4.cell(row=row, column=2, value=f.get("severity", "")).border = THIN_BORDER
        ws4.cell(row=row, column=3, value=sanitize_string(f.get("title", ""))).border = THIN_BORDER
        ws4.cell(row=row, column=4, value=sanitize_string(f.get("where", ""))).border = THIN_BORDER
        ws4.cell(row=row, column=5, value=sanitize_string(f.get("note", ""))).border = THIN_BORDER
        ws4.cell(row=row, column=6, value="待确认").border = THIN_BORDER
        row += 1

    # Sheet 5: Egress Control
    ws5 = wb.create_sheet("出口网络风险")

    egress_path = SCAN_DIR / "egress-control-audit/result.json"
    egress_findings = []

    if egress_path.exists():
        data = load_json(egress_path)
        egress_findings = data.get("findings", [])

    headers5 = ["序号", "严重度", "类型", "位置", "说明", "确认状态"]
    for col, header in enumerate(headers5, 1):
        cell = ws5.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    row = 2
    for i, f in enumerate(egress_findings, 1):
        ws5.cell(row=row, column=1, value=i).border = THIN_BORDER
        ws5.cell(row=row, column=2, value=f.get("severity", "")).border = THIN_BORDER
        ws5.cell(row=row, column=3, value=sanitize_string(f.get("title", ""))).border = THIN_BORDER
        ws5.cell(row=row, column=4, value=sanitize_string(f.get("where", ""))).border = THIN_BORDER
        ws5.cell(row=row, column=5, value=sanitize_string(f.get("note", ""))).border = THIN_BORDER
        ws5.cell(row=row, column=6, value="待确认").border = THIN_BORDER
        row += 1

    # Sheet 6: Summary
    ws6 = wb.create_sheet("汇总统计")

    summary_data = [
        ["模块", "总数", "Critical", "High", "Medium", "Low", "误报数"],
    ]

    # Calculate stats
    gitleaks_fp = sum(1 for f in gitleaks_findings if is_false_positive(f, "gitleaks"))

    # Calculate severity for gitleaks (it doesn't have severity field)
    gitleaks_critical = 0
    gitleaks_high = 0
    gitleaks_medium = 0
    gitleaks_low = 0

    for f in gitleaks_findings:
        fp = is_false_positive(f, "gitleaks")
        if fp:
            continue
        entropy = f.get("Entropy", 0)
        if entropy >= 4.5:
            gitleaks_high += 1
        elif entropy >= 3.5:
            gitleaks_medium += 1
        else:
            gitleaks_low += 1

    summary_data.append(["Gitleaks敏感信息", len(gitleaks_findings),
        gitleaks_critical, gitleaks_high, gitleaks_medium, gitleaks_low,
        gitleaks_fp])

    summary_data.append(["TruffleHog密钥", len(trufflehog_findings), "-", "-", "-", "-",
        sum(1 for f in trufflehog_findings if is_false_positive(f, "trufflehog"))])

    summary_data.append(["提权逃逸检查", len(privesc_findings),
        sum(1 for f in privesc_findings if f.get("severity")=="critical"),
        sum(1 for f in privesc_findings if f.get("severity")=="high"),
        sum(1 for f in privesc_findings if f.get("severity")=="medium"),
        sum(1 for f in privesc_findings if f.get("severity")=="low"), 0])

    summary_data.append(["横向移动风险", len(lateral_findings),
        sum(1 for f in lateral_findings if f.get("severity")=="critical"),
        sum(1 for f in lateral_findings if f.get("severity")=="high"),
        sum(1 for f in lateral_findings if f.get("severity")=="medium"),
        sum(1 for f in lateral_findings if f.get("severity")=="low"), 0])

    summary_data.append(["出口网络风险", len(egress_findings),
        sum(1 for f in egress_findings if f.get("severity")=="critical"),
        sum(1 for f in egress_findings if f.get("severity")=="high"),
        sum(1 for f in egress_findings if f.get("severity")=="medium"),
        sum(1 for f in egress_findings if f.get("severity")=="low"), 0])

    for row_idx, row_data in enumerate(summary_data, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws6.cell(row=row_idx, column=col_idx, value=value)
            cell.border = THIN_BORDER
            if row_idx == 1:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT

    # Save
    output_path = OUTPUT_DIR / "detailed-audit-findings.xlsx"
    wb.save(output_path)
    print(f"Excel report saved to: {output_path}")
    print(f"  - Gitleaks: {len(gitleaks_findings)} findings")
    print(f"  - TruffleHog: {len(trufflehog_findings)} findings")
    print(f"  - Privilege Escalation: {len(privesc_findings)} findings")
    print(f"  - Lateral Movement: {len(lateral_findings)} findings")
    print(f"  - Egress Control: {len(egress_findings)} findings")

if __name__ == "__main__":
    # Support command line argument for scan directory
    if len(sys.argv) > 1:
        scan_dir = Path(sys.argv[1])
        output_dir = scan_dir.parent
        create_excel_report(scan_dir, output_dir)
    else:
        create_excel_report()
