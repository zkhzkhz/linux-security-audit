#!/usr/bin/env python3
"""
Audit Report Analyzer - Main analysis script.
Takes a compressed audit report archive and generates comprehensive reports.
"""
import json
import sys
import os
import argparse
import shutil
import tarfile
import zipfile
import tempfile
from pathlib import Path
from datetime import datetime
from collections import defaultdict, Counter

# Try to import openpyxl for Excel generation
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not installed, Excel report will be skipped")

# Styles for Excel
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CRITICAL_FILL = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
HIGH_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
MEDIUM_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LOW_FILL = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
FP_FILL = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def log(msg, level="INFO"):
    """Print log message with timestamp."""
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    print(f"[{level} {timestamp}] {msg}")


def extract_archive(archive_path, output_dir):
    """Extract archive to output directory."""
    archive_path = Path(archive_path)
    output_dir = Path(output_dir)

    # Handle nested archives (tar.zip -> tar.gz)
    temp_dir = tempfile.mkdtemp()

    if archive_path.suffix == '.zip':
        with zipfile.ZipFile(archive_path, 'r') as zf:
            zf.extractall(temp_dir)
        # Check for nested tar.gz
        for f in Path(temp_dir).iterdir():
            if f.suffix == '.gz' or f.name.endswith('.tar.gz'):
                archive_path = f
                break
    else:
        shutil.copy(archive_path, temp_dir)
        archive_path = Path(temp_dir) / archive_path.name

    # Extract tar.gz
    if archive_path.name.endswith('.tar.gz') or archive_path.suffix == '.gz':
        with tarfile.open(archive_path, 'r:gz') as tf:
            tf.extractall(output_dir)
    elif archive_path.suffix == '.zip':
        with zipfile.ZipFile(archive_path, 'r') as zf:
            zf.extractall(output_dir)
    else:
        raise ValueError(f"Unsupported archive format: {archive_path}")

    shutil.rmtree(temp_dir, ignore_errors=True)
    log(f"Extracted archive to: {output_dir}")


def find_scan_dir(output_dir):
    """Find the scan results directory."""
    output_dir = Path(output_dir)

    # Look for summary.json or result.json
    for summary in output_dir.rglob("summary.json"):
        return summary.parent

    for result in output_dir.rglob("sensitive-info-scan/result.json"):
        return result.parent.parent

    # Return the first directory with reports
    for d in output_dir.rglob("reports/*"):
        if d.is_dir():
            return d

    return output_dir


def load_json(path):
    """Load JSON file safely."""
    path = Path(path)
    if not path.exists():
        return {}
    try:
        with open(path, 'r', encoding='utf-8', errors='replace') as f:
            return json.load(f)
    except Exception as e:
        log(f"Error loading {path}: {e}", "WARN")
        return {}


def load_json_lines(path):
    """Load JSONL file (one JSON per line)."""
    results = []
    path = Path(path)
    if not path.exists():
        return results
    try:
        with open(path, 'r', encoding='utf-8', errors='replace') as f:
            for line in f:
                line = line.strip()
                if line and line != '[]':
                    try:
                        results.append(json.loads(line))
                    except:
                        pass
    except Exception as e:
        log(f"Error loading {path}: {e}", "WARN")
    return results


def is_false_positive(finding, source="gitleaks"):
    """Determine if a finding is likely a false positive."""
    where = finding.get("File", "") or finding.get("where", "")
    if not where:
        where = finding.get("SourceMetadata", {}).get("Data", {}).get("Filesystem", {}).get("file", "")

    note = finding.get("note", "") or finding.get("Description", "")
    title = finding.get("title", "") or finding.get("DetectorName", "") or finding.get("RuleID", "")

    # FP patterns
    if "/overlay2/" in where:
        if any(x in where for x in ["/usr/share/doc/", "/usr/share/misc/", "/usr/lib/", "/var/cache/dnf/"]):
            return "Docker镜像层系统文件"
        if "/usr/local/go/" in where and ("testdata" in where or "test" in where):
            return "Docker镜像层Go测试数据"
        if "/home/" in where and "/lib/python" in where:
            return "Docker镜像层Python stdlib"

    if "/lib/python" in where or "/lib64/python" in where:
        if any(x in where for x in ["urllib/request.py", "distutils/", "nntplib.py", "hashlib.py"]):
            return "Python stdlib源码误匹配"
        if "/test/" in where:
            return "Python stdlib测试代码"

    if any(where.endswith(ext) for ext in [".mgc", ".solv", ".db", ".bolt", ".bin"]):
        return "二进制文件误匹配"

    if "/test/" in where or "/tests/" in where or "/testdata/" in where:
        return "测试数据"

    if "/usr/share/doc/" in where:
        return "文档文件"

    # JDK Eclipse files
    if "ECLIPSE_.RSA" in where or "META-INF" in where:
        return "JDK签名文件误匹配"

    if finding.get("is_likely_fp", False):
        return "自动识别误报"

    entropy = finding.get("Entropy", 0)
    if entropy and entropy < 3.0:
        return "低熵值"

    return None


def sanitize_string(s, max_len=5000):
    """Remove illegal characters for Excel."""
    if not s:
        return ""
    s = str(s)[:max_len]
    illegal_chars = []
    for i in range(0, 32):
        if i not in (9, 10, 13):
            illegal_chars.append(chr(i))
    for char in illegal_chars:
        s = s.replace(char, '')
    return s


def analyze_kernel_cves(privesc_dir):
    """Analyze kernel vulnerabilities from LinPEAS output."""
    linpeas_file = privesc_dir / "linpeas_raw.txt"
    if not linpeas_file.exists():
        return []

    cves = []
    content = linpeas_file.read_text(encoding='utf-8', errors='replace')

    # CVE-2026-31431 (Copy Fail)
    if "CVE-2026-31431" in content:
        if "NOT VULNERABLE" in content:
            cves.append({
                "cve": "CVE-2026-31431",
                "name": "Copy Fail",
                "status": "not_affected",
                "severity": "medium"
            })
        else:
            cves.append({
                "cve": "CVE-2026-31431",
                "name": "Copy Fail",
                "status": "potentially_affected",
                "severity": "medium"
            })

    # CVE-2026-43284/43500 (Dirty Frag)
    if "CVE-2026-43284" in content or "CVE-2026-43500" in content:
        if "LIKELY VULNERABLE" in content:
            cves.append({
                "cve": "CVE-2026-43284",
                "name": "Dirty Frag (xfrm-ESP)",
                "status": "likely_affected",
                "severity": "medium"
            })
            cves.append({
                "cve": "CVE-2026-43500",
                "name": "Dirty Frag (rxrpc)",
                "status": "likely_affected",
                "severity": "medium"
            })

    # CVE-2026-41651 (Pack2TheRoot)
    if "CVE-2026-41651" in content:
        if "Vulnerable to CVE-2026-41651" in content:
            cves.append({
                "cve": "CVE-2026-41651",
                "name": "Pack2TheRoot",
                "status": "affected",
                "severity": "medium"
            })

    return cves


def analyze_cdk_results(privesc_dir):
    """Analyze CDK container escape tool results."""
    cdk_dir = privesc_dir / "cdk"
    results = []

    if not cdk_dir.exists():
        return results

    # Look for cdk.log
    cdk_log = cdk_dir / "cdk.log"
    if cdk_log.exists():
        content = cdk_log.read_text(encoding='utf-8', errors='replace')
        # Parse CDK findings
        for line in content.split('\n'):
            if 'VULNERABLE' in line.upper() or 'EXPLOIT' in line.upper():
                results.append({
                    "source": "cdk",
                    "type": "container_escape",
                    "finding": line[:200]
                })

    # Check individual container results
    for container_dir in cdk_dir.iterdir():
        if container_dir.is_dir() and container_dir.name.startswith("k8s_"):
            result_file = container_dir / "result.json"
            if result_file.exists():
                try:
                    data = json.loads(result_file.read_text())
                    if data.get("findings"):
                        results.append({
                            "source": "cdk",
                            "container": container_dir.name,
                            "findings": data["findings"]
                        })
                except:
                    pass

    return results


def analyze_deepce_results(privesc_dir):
    """Analyze DeepCE container escape results."""
    deepce_dir = privesc_dir / "deepce"
    results = []

    if not deepce_dir.exists():
        return results

    result_file = deepce_dir / "result.json"
    if result_file.exists():
        try:
            data = json.loads(result_file.read_text())
            results = data.get("findings", [])
        except:
            pass

    return results


def analyze_peirates_results(privesc_dir):
    """Analyze Peirates K8s privilege escalation results."""
    peirates_dir = privesc_dir / "peirates"
    results = []

    if not peirates_dir.exists():
        return results

    result_file = peirates_dir / "result.json"
    if result_file.exists():
        try:
            data = json.loads(result_file.read_text())
            results = data.get("findings", [])
        except:
            pass

    return results


def analyze_kubebench_results(privesc_dir):
    """Analyze Kube-bench CIS benchmark results."""
    kubebench_dir = privesc_dir / "kube-bench"
    results = []

    if not kubebench_dir.exists():
        return results

    result_file = kubebench_dir / "result.json"
    if result_file.exists():
        try:
            data = json.loads(result_file.read_text())
            # Extract failed checks
            for test in data.get("Tests", []):
                for res in test.get("results", []):
                    if res.get("status") == "FAIL":
                        results.append({
                            "id": res.get("test_number", ""),
                            "desc": res.get("test_desc", ""),
                            "remediation": res.get("remediation", "")
                        })
        except:
            pass

    return results


def analyze_trufflehog_results(sensitive_dir):
    """Analyze TruffleHog secret scan results."""
    trufflehog_dir = sensitive_dir / "trufflehog"
    results = []

    if not trufflehog_dir.exists():
        return results

    for f in trufflehog_dir.glob("*.json"):
        try:
            content = f.read_text()
            for line in content.split('\n'):
                if line.strip():
                    try:
                        data = json.loads(line)
                        results.append(data)
                    except:
                        pass
        except:
            pass

    return results


def analyze_lateral_movement(lateral_dir):
    """Analyze lateral movement scan results."""
    result_file = lateral_dir / "result.json"
    if result_file.exists():
        return load_json(result_file).get("findings", [])
    return []


def analyze_egress_control(egress_dir):
    """Analyze egress control audit results."""
    result_file = egress_dir / "result.json"
    if result_file.exists():
        return load_json(result_file).get("findings", [])
    return []


def generate_excel_report(scan_dir, output_dir, scan_data):
    """Generate Excel report with all findings."""
    if not HAS_OPENPYXL:
        log("openpyxl not available, skipping Excel report", "WARN")
        return None

    wb = openpyxl.Workbook()

    # Sheet 1: Gitleaks findings
    ws1 = wb.active
    ws1.title = "Gitleaks敏感信息"

    gitleaks_findings = scan_data.get("sensitive_raw", [])

    headers = ["序号", "严重度", "类型", "文件路径", "行号", "密钥前缀", "完整密钥", "评分信息", "误报原因", "确认状态"]
    for col, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    row = 2
    for i, f in enumerate(gitleaks_findings[:5000], 1):
        file_path = f.get("File", "") or f.get("where", "")
        line_num = f.get("StartLine", "") or ""
        if not line_num and ":" in file_path:
            parts = file_path.rsplit(":", 1)
            file_path = parts[0]
            line_num = parts[1]

        secret = f.get("Secret", "") or f.get("secret_prefix", "") or f.get("Match", "")
        title = f.get("RuleID", "") or f.get("title", "")
        severity = f.get("severity", "")
        if not severity:
            entropy = f.get("Entropy", 0)
            if entropy >= 4.5:
                severity = "high"
            elif entropy >= 3.5:
                severity = "medium"
            else:
                severity = "low"

        note = f.get("note", "") or f.get("Description", "")
        entropy_str = f"entropy={f.get('Entropy', 0):.2f}" if f.get("Entropy") else ""
        fp_reason = is_false_positive(f, "gitleaks")

        ws1.cell(row=row, column=1, value=i).border = THIN_BORDER
        ws1.cell(row=row, column=2, value=severity).border = THIN_BORDER
        ws1.cell(row=row, column=3, value=sanitize_string(title)).border = THIN_BORDER
        ws1.cell(row=row, column=4, value=sanitize_string(file_path)).border = THIN_BORDER
        ws1.cell(row=row, column=5, value=line_num).border = THIN_BORDER
        ws1.cell(row=row, column=6, value=sanitize_string(secret[:50] if secret else "")).border = THIN_BORDER
        ws1.cell(row=row, column=7, value=sanitize_string(secret)).border = THIN_BORDER
        ws1.cell(row=row, column=8, value=sanitize_string(f"{note} {entropy_str}".strip()[:100])).border = THIN_BORDER
        ws1.cell(row=row, column=9, value=sanitize_string(fp_reason or "")).border = THIN_BORDER
        ws1.cell(row=row, column=10, value="待确认").border = THIN_BORDER

        if fp_reason:
            for col in range(1, 11):
                ws1.cell(row=row, column=col).fill = FP_FILL
        elif severity == "critical":
            for col in range(1, 11):
                ws1.cell(row=row, column=col).fill = CRITICAL_FILL
        elif severity == "high":
            for col in range(1, 11):
                ws1.cell(row=row, column=col).fill = HIGH_FILL
        elif severity == "medium":
            for col in range(1, 11):
                ws1.cell(row=row, column=col).fill = MEDIUM_FILL

        row += 1

    # Adjust column widths
    ws1.column_dimensions['A'].width = 8
    ws1.column_dimensions['B'].width = 12
    ws1.column_dimensions['C'].width = 25
    ws1.column_dimensions['D'].width = 60
    ws1.column_dimensions['E'].width = 8
    ws1.column_dimensions['F'].width = 30
    ws1.column_dimensions['G'].width = 80
    ws1.column_dimensions['H'].width = 40
    ws1.column_dimensions['I'].width = 20
    ws1.column_dimensions['J'].width = 12

    # Sheet 2: Privilege Escalation
    ws2 = wb.create_sheet("提权逃逸检查")
    privesc_findings = scan_data.get("privesc", {}).get("findings", [])

    headers2 = ["序号", "严重度", "类型", "位置", "说明", "确认状态"]
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = THIN_BORDER

    row = 2
    for i, f in enumerate(privesc_findings, 1):
        severity = f.get("severity", "")
        ws2.cell(row=row, column=1, value=i).border = THIN_BORDER
        ws2.cell(row=row, column=2, value=severity).border = THIN_BORDER
        ws2.cell(row=row, column=3, value=sanitize_string(f.get("title", ""))).border = THIN_BORDER
        ws2.cell(row=row, column=4, value=sanitize_string(f.get("where", ""))).border = THIN_BORDER
        ws2.cell(row=row, column=5, value=sanitize_string(f.get("note", "")[:100])).border = THIN_BORDER
        ws2.cell(row=row, column=6, value="待确认").border = THIN_BORDER

        if severity == "critical":
            for col in range(1, 7):
                ws2.cell(row=row, column=col).fill = CRITICAL_FILL
        elif severity == "high":
            for col in range(1, 7):
                ws2.cell(row=row, column=col).fill = HIGH_FILL

        row += 1

    # Sheet 3: Summary
    ws3 = wb.create_sheet("汇总统计")
    summary_data = scan_data.get("summary", {})

    ws3.cell(row=1, column=1, value="模块").border = THIN_BORDER
    ws3.cell(row=1, column=2, value="总数").border = THIN_BORDER
    ws3.cell(row=1, column=3, value="Critical").border = THIN_BORDER
    ws3.cell(row=1, column=4, value="High").border = THIN_BORDER
    ws3.cell(row=1, column=5, value="Medium").border = THIN_BORDER
    ws3.cell(row=1, column=6, value="Low").border = THIN_BORDER
    ws3.cell(row=1, column=7, value="误报数").border = THIN_BORDER

    for col in range(1, 8):
        ws3.cell(row=1, column=col).fill = HEADER_FILL
        ws3.cell(row=1, column=col).font = HEADER_FONT

    row = 2
    for module, counts in summary_data.items():
        ws3.cell(row=row, column=1, value=module).border = THIN_BORDER
        ws3.cell(row=row, column=2, value=counts.get("total", 0)).border = THIN_BORDER
        ws3.cell(row=row, column=3, value=counts.get("critical", 0)).border = THIN_BORDER
        ws3.cell(row=row, column=4, value=counts.get("high", 0)).border = THIN_BORDER
        ws3.cell(row=row, column=5, value=counts.get("medium", 0)).border = THIN_BORDER
        ws3.cell(row=row, column=6, value=counts.get("low", 0)).border = THIN_BORDER
        ws3.cell(row=row, column=7, value=counts.get("fp", 0)).border = THIN_BORDER
        row += 1

    # Save
    output_path = output_dir / "detailed-audit-findings.xlsx"
    wb.save(output_path)
    return output_path


def generate_evidence_report(scan_dir, output_dir, scan_data):
    """Generate detailed evidence report in Markdown."""
    lines = []

    host_info = scan_data.get("host", {})
    summary = scan_data.get("summary", {})

    lines.append("# 安全审计证据报告")
    lines.append("")
    lines.append(f"**生成时间：** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"**目标主机：** {host_info.get('host', 'unknown')} ({host_info.get('arch', 'unknown')})")
    lines.append(f"**扫描目录：** {scan_dir}")
    lines.append("")
    lines.append("---")
    lines.append("")

    # Summary table
    lines.append("## 一、扫描结果总览")
    lines.append("")
    lines.append("| 模块 | 状态 | Critical | High | Medium | Low | 误报数 |")
    lines.append("|------|------|----------|------|--------|-----|--------|")

    for module, counts in summary.items():
        lines.append(f"| {module} | {counts.get('status', 'unknown')} | {counts.get('critical', 0)} | {counts.get('high', 0)} | {counts.get('medium', 0)} | {counts.get('low', 0)} | {counts.get('fp', 0)} |")

    lines.append("")

    # Critical findings
    lines.append("## 二、Critical级别发现（含证据）")
    lines.append("")

    critical_findings = [f for f in scan_data.get("privesc", {}).get("findings", []) if f.get("severity") == "critical"]
    lines.append(f"共发现 **{len(critical_findings)}** 个 Critical 级别问题")
    lines.append("")

    for i, f in enumerate(critical_findings, 1):
        lines.append(f"### 2.{i} {f.get('title', 'unknown')}")
        lines.append("")
        lines.append(f"- **严重度：** {f.get('severity', 'unknown')}")
        lines.append(f"- **位置：** `{f.get('where', 'unknown')}`")
        lines.append(f"- **说明：** {f.get('note', 'N/A')[:200]}")
        lines.append("")

    # High findings
    lines.append("## 三、High级别发现（含证据）")
    lines.append("")

    high_findings = [f for f in scan_data.get("privesc", {}).get("findings", []) if f.get("severity") == "high"]
    lines.append(f"共发现 **{len(high_findings)}** 个 High 级别问题")
    lines.append("")

    # Group by type
    high_by_type = defaultdict(list)
    for f in high_findings:
        high_by_type[f.get("title", "other")].append(f)

    lines.append("| 类型 | 数量 | 示例位置 |")
    lines.append("|------|------|----------|")

    for title, items in sorted(high_by_type.items(), key=lambda x: -len(x[1])):
        example = items[0].get("where", "")[:50]
        lines.append(f"| {title} | {len(items)} | `{example}...` |")

    lines.append("")

    # Kernel CVEs
    kernel_cves = scan_data.get("kernel_cves", [])
    if kernel_cves:
        lines.append("## 四、Kernel漏洞发现")
        lines.append("")
        lines.append("| CVE | 名称 | 状态 | 严重度 |")
        lines.append("|-----|------|------|--------|")

        for cve in kernel_cves:
            lines.append(f"| {cve['cve']} | {cve['name']} | {cve['status']} | {cve['severity']} |")

        lines.append("")
        lines.append("**详细分析见：** kernel-vulnerability-analysis.md")
        lines.append("")

    # False positive analysis
    lines.append("## 五、误报分析")
    lines.append("")

    fp_reasons = Counter()
    for f in scan_data.get("sensitive_raw", []):
        fp = is_false_positive(f, "gitleaks")
        if fp:
            fp_reasons[fp] += 1

    if fp_reasons:
        lines.append("| 误报原因 | 数量 |")
        lines.append("|----------|------|")
        for reason, count in fp_reasons.most_common():
            lines.append(f"| {reason} | {count} |")
        lines.append("")

    # Remediation
    lines.append("## 六、其他工具扫描结果")
    lines.append("")

    # CDK results
    cdk_results = scan_data.get("cdk", [])
    if cdk_results:
        lines.append("### 6.1 CDK容器逃逸检测")
        lines.append("")
        lines.append(f"共发现 **{len(cdk_results)}** 个容器逃逸风险")
        lines.append("")
        lines.append("| 容器 | 发现类型 |")
        lines.append("|------|----------|")
        for r in cdk_results[:20]:
            container = r.get("container", "unknown")[:40]
            finding = r.get("finding", r.get("type", ""))[:50]
            lines.append(f"| `{container}...` | {finding} |")
        lines.append("")

    # DeepCE results
    deepce_results = scan_data.get("deepce", [])
    if deepce_results:
        lines.append("### 6.2 DeepCE容器检测")
        lines.append("")
        lines.append(f"共发现 **{len(deepce_results)}** 个风险")
        lines.append("")
        lines.append("| 类型 | 数量 |")
        lines.append("|------|------|")
        by_type = Counter(r.get("title", "unknown") for r in deepce_results)
        for t, c in by_type.most_common(10):
            lines.append(f"| {t} | {c} |")
        lines.append("")

    # Peirates results
    peirates_results = scan_data.get("peirates", [])
    if peirates_results:
        lines.append("### 6.3 Peirates K8s提权检测")
        lines.append("")
        lines.append(f"共发现 **{len(peirates_results)}** 个K8s提权风险")
        lines.append("")

    # Kube-bench results
    kubebench_results = scan_data.get("kubebench", [])
    if kubebench_results:
        lines.append("### 6.4 Kube-bench CIS基准检查")
        lines.append("")
        lines.append(f"共 **{len(kubebench_results)}** 项未通过")
        lines.append("")
        lines.append("| 检查项 | 描述 |")
        lines.append("|--------|------|")
        for r in kubebench_results[:15]:
            lines.append(f"| {r.get('id', '')} | {r.get('desc', '')[:60]} |")
        lines.append("")

    # TruffleHog results
    trufflehog_results = scan_data.get("trufflehog", [])
    if trufflehog_results:
        lines.append("### 6.5 TruffleHog密钥检测")
        lines.append("")
        lines.append(f"共发现 **{len(trufflehog_results)}** 个密钥")
        lines.append("")
        lines.append("| 检测器 | 文件 |")
        lines.append("|--------|------|")
        for r in trufflehog_results[:15]:
            detector = r.get("DetectorName", "unknown")
            file_path = r.get("SourceMetadata", {}).get("Data", {}).get("Filesystem", {}).get("file", "")[:50]
            lines.append(f"| {detector} | `{file_path}...` |")
        lines.append("")

    # Lateral movement
    lateral_findings = scan_data.get("lateral_findings", [])
    if lateral_findings:
        lines.append("### 6.6 横向移动风险")
        lines.append("")
        lines.append(f"共发现 **{len(lateral_findings)}** 个横向移动风险")
        lines.append("")

    # Egress control
    egress_findings = scan_data.get("egress_findings", [])
    if egress_findings:
        lines.append("### 6.7 出口网络风险")
        lines.append("")
        lines.append(f"共发现 **{len(egress_findings)}** 个出口风险")
        lines.append("")

    # Remediation
    lines.append("## 七、修复建议")
    lines.append("")
    lines.append("### P0 - 立即处理")
    lines.append("")
    lines.append("| 问题 | 修复建议 |")
    lines.append("|------|----------|")
    lines.append("| Docker Socket可写 | 移除Docker组权限或限制socket权限 |")
    lines.append("| Cron目录可写 | chmod 755 /etc/cron.* |")
    lines.append("| PackageKit漏洞 | apt upgrade packagekit |")
    lines.append("")
    lines.append("### P1 - 高优先级")
    lines.append("")
    lines.append("| 问题 | 修复建议 |")
    lines.append("|------|----------|")
    lines.append("| Dirty Frag漏洞 | 升级内核或禁用危险模块 |")
    lines.append("| 容器挂载风险 | 改为只读或移除挂载 |")
    lines.append("| K8s CIS不合规 | 根据Kube-bench建议修复 |")
    lines.append("")

    # Append raw LinPEAS evidence for critical findings
    lines.append("## 八、关键发现原始证据")
    lines.append("")

    # LinPEAS critical sections
    privesc_dir = scan_dir / "privesc-escape-check"
    linpeas_file = privesc_dir / "linpeas_raw.txt"
    if linpeas_file.exists():
        linpeas_content = linpeas_file.read_text(encoding='utf-8', errors='replace')

        # Extract Docker socket section
        if "docker.sock" in linpeas_content.lower():
            lines.append("### 8.1 Docker Socket检测证据")
            lines.append("")
            lines.append("```")
            # Find the section
            idx = linpeas_content.lower().find("docker.sock")
            start = max(0, idx - 200)
            end = min(len(linpeas_content), idx + 500)
            lines.append(linpeas_content[start:end])
            lines.append("```")
            lines.append("")

        # Extract Cron section
        if "cron" in linpeas_content.lower() and "writable" in linpeas_content.lower():
            lines.append("### 8.2 Cron目录检测证据")
            lines.append("")
            lines.append("```")
            idx = linpeas_content.lower().find("writable cron")
            if idx > 0:
                start = max(0, idx - 100)
                end = min(len(linpeas_content), idx + 800)
                lines.append(linpeas_content[start:end])
            lines.append("```")
            lines.append("")

    # CDK log evidence
    cdk_log = privesc_dir / "cdk" / "cdk.log"
    if cdk_log.exists():
        lines.append("### 8.3 CDK检测证据")
        lines.append("")
        lines.append("```")
        content = cdk_log.read_text(encoding='utf-8', errors='replace')
        # Show first 2000 chars of relevant findings
        lines.append(content[:2000] if len(content) > 2000 else content)
        lines.append("```")
        lines.append("")
    lines.append("")
    lines.append("| 问题 | 修复建议 |")
    lines.append("|------|----------|")
    lines.append("| Dirty Frag漏洞 | 升级内核或禁用危险模块 |")
    lines.append("| 容器挂载风险 | 改为只读或移除挂载 |")
    lines.append("")

    output_path = output_dir / "detailed-audit-evidence.md"
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    return output_path


def generate_kernel_report(scan_dir, output_dir, scan_data):
    """Generate kernel vulnerability analysis report."""
    kernel_cves = scan_data.get("kernel_cves", [])
    if not kernel_cves:
        return None

    linpeas_file = scan_dir / "privesc-escape-check/linpeas_raw.txt"
    linpeas_content = ""
    if linpeas_file.exists():
        linpeas_content = linpeas_file.read_text(encoding='utf-8', errors='replace')

    lines = []
    lines.append("# Kernel漏洞详细分析报告")
    lines.append("")
    lines.append(f"**生成时间：** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append(f"**目标主机：** {scan_data.get('host', {}).get('host', 'unknown')}")
    lines.append("")
    lines.append("---")
    lines.append("")

    for cve in kernel_cves:
        lines.append(f"## {cve['cve']} ({cve['name']})")
        lines.append("")
        lines.append(f"**状态：** {cve['status']}")
        lines.append(f"**严重度：** {cve['severity']}")
        lines.append("")

        # Extract relevant LinPEAS output
        if cve['cve'] in linpeas_content:
            lines.append("**LinPEAS扫描证据：**")
            lines.append("```")
            # Find the section
            start = linpeas_content.find(cve['cve'])
            if start > 0:
                # Find section start
                section_start = linpeas_content.rfind("╔", 0, start)
                if section_start > 0:
                    section_end = linpeas_content.find("\n\n", section_start)
                    if section_end > 0:
                        lines.append(linpeas_content[section_start:section_end])
            lines.append("```")
            lines.append("")

        # Add remediation based on CVE
        if cve['cve'] == "CVE-2026-41651":
            lines.append("**修复建议：**")
            lines.append("```bash")
            lines.append("apt update && apt upgrade packagekit")
            lines.append("```")
        elif cve['cve'] in ["CVE-2026-43284", "CVE-2026-43500"]:
            lines.append("**修复建议：**")
            lines.append("```bash")
            lines.append("# 禁用危险模块")
            lines.append("echo 'install esp4 /bin/false' >> /etc/modprobe.d/disable-dirtyfrag.conf")
            lines.append("echo 'install esp6 /bin/false' >> /etc/modprobe.d/disable-dirtyfrag.conf")
            lines.append("echo 'install rxrpc /bin/false' >> /etc/modprobe.d/disable-dirtyfrag.conf")
            lines.append("# 或升级内核")
            lines.append("apt upgrade linux-image-generic")
            lines.append("```")

        lines.append("")

    output_path = output_dir / "kernel-vulnerability-analysis.md"
    with open(output_path, 'w', encoding='utf-8') as f:
        f.write('\n'.join(lines))

    return output_path


def generate_json_summary(output_dir, scan_data):
    """Generate JSON summary."""
    summary = {
        "generated": datetime.now().isoformat(),
        "host": scan_data.get("host", {}),
        "summary": scan_data.get("summary", {}),
        "kernel_cves": scan_data.get("kernel_cves", [])
    }

    output_path = output_dir / "audit-summary.json"
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(summary, f, ensure_ascii=False, indent=2)

    return output_path


def analyze_report(archive_path, output_base=None):
    """Main analysis function."""
    archive_path = Path(archive_path)

    if not archive_path.exists():
        raise FileNotFoundError(f"Archive not found: {archive_path}")

    # Create output directory
    if output_base:
        output_base = Path(output_base)
    else:
        output_base = Path("/root/linux-security-audit/reports")

    # Generate unique directory name
    report_name = archive_path.stem.replace('.tar', '').replace('.zip', '')
    output_dir = output_base / f"{report_name}-analysis"
    output_dir.mkdir(parents=True, exist_ok=True)

    log(f"Output directory: {output_dir}")

    # Extract archive
    log("Extracting archive...")
    extract_archive(archive_path, output_dir)

    # Find scan directory
    scan_dir = find_scan_dir(output_dir)
    log(f"Scan directory: {scan_dir}")

    # Load all data
    log("Loading scan results...")

    # Load summary.json
    summary_json = load_json(scan_dir.parent / "summary.json")

    # Load sensitive info scan
    sensitive_result = load_json(scan_dir / "sensitive-info-scan/result.json")
    sensitive_raw = load_json(scan_dir / "sensitive-info-scan/raw.json")
    if not isinstance(sensitive_raw, list):
        sensitive_raw = []

    # Load privesc
    privesc_result = load_json(scan_dir / "privesc-escape-check/result.json")

    # Load lateral movement
    lateral_result = load_json(scan_dir / "lateral-movement-scan/result.json")

    # Load egress control
    egress_result = load_json(scan_dir / "egress-control-audit/result.json")

    # Analyze kernel CVEs
    kernel_cves = analyze_kernel_cves(scan_dir / "privesc-escape-check")

    # Analyze other tools
    log("Analyzing CDK results...")
    cdk_results = analyze_cdk_results(scan_dir / "privesc-escape-check")

    log("Analyzing DeepCE results...")
    deepce_results = analyze_deepce_results(scan_dir / "privesc-escape-check")

    log("Analyzing Peirates results...")
    peirates_results = analyze_peirates_results(scan_dir / "privesc-escape-check")

    log("Analyzing Kube-bench results...")
    kubebench_results = analyze_kubebench_results(scan_dir / "privesc-escape-check")

    log("Analyzing TruffleHog results...")
    trufflehog_results = analyze_trufflehog_results(scan_dir / "sensitive-info-scan")

    log("Analyzing lateral movement...")
    lateral_findings = analyze_lateral_movement(scan_dir / "lateral-movement-scan")

    log("Analyzing egress control...")
    egress_findings = analyze_egress_control(scan_dir / "egress-control-audit")

    # Build scan data
    scan_data = {
        "host": {
            "host": summary_json.get("host", "unknown"),
            "arch": summary_json.get("arch", "unknown")
        },
        "sensitive": sensitive_result,
        "sensitive_raw": sensitive_raw,
        "privesc": privesc_result,
        "lateral": lateral_result,
        "egress": egress_result,
        "kernel_cves": kernel_cves,
        # Additional tool results
        "cdk": cdk_results,
        "deepce": deepce_results,
        "peirates": peirates_results,
        "kubebench": kubebench_results,
        "trufflehog": trufflehog_results,
        "lateral_findings": lateral_findings,
        "egress_findings": egress_findings,
        "summary": {}
    }

    # Build summary counts
    for name, result, fp_key in [
        ("敏感信息扫描", sensitive_result, "likely_fp"),
        ("提权逃逸检查", privesc_result, None),
        ("横向移动扫描", lateral_result, None),
        ("出口控制审计", egress_result, None),
    ]:
        counts = result.get("counts", {})
        scan_data["summary"][name] = {
            "status": result.get("status", "unknown"),
            "total": len(result.get("findings", [])),
            "critical": counts.get("critical", 0),
            "high": counts.get("high", 0),
            "medium": counts.get("medium", 0),
            "low": counts.get("low", 0),
            "fp": counts.get(fp_key, 0) if fp_key else 0
        }

    # Add tool-specific summaries
    scan_data["summary"]["CDK容器逃逸"] = {
        "status": "ok" if cdk_results else "skip",
        "total": len(cdk_results),
        "high": sum(1 for r in cdk_results if r.get("severity") == "high")
    }
    scan_data["summary"]["DeepCE检测"] = {
        "status": "ok" if deepce_results else "skip",
        "total": len(deepce_results),
        "critical": sum(1 for r in deepce_results if r.get("severity") == "critical"),
        "high": sum(1 for r in deepce_results if r.get("severity") == "high")
    }
    scan_data["summary"]["Peirates K8s提权"] = {
        "status": "ok" if peirates_results else "skip",
        "total": len(peirates_results),
        "high": sum(1 for r in peirates_results if r.get("severity") == "high")
    }
    scan_data["summary"]["Kube-bench CIS"] = {
        "status": "ok" if kubebench_results else "skip",
        "total": len(kubebench_results),
        "high": len(kubebench_results)  # All failed checks are high
    }
    scan_data["summary"]["TruffleHog密钥"] = {
        "status": "ok" if trufflehog_results else "skip",
        "total": len(trufflehog_results)
    }

    # Generate reports
    log("Generating reports...")

    # 1. Excel report
    excel_path = generate_excel_report(scan_dir, output_dir, scan_data)
    if excel_path:
        log(f"Excel report: {excel_path}")

    # 2. Evidence report
    evidence_path = generate_evidence_report(scan_dir, output_dir, scan_data)
    log(f"Evidence report: {evidence_path}")

    # 3. Kernel report
    kernel_path = generate_kernel_report(scan_dir, output_dir, scan_data)
    if kernel_path:
        log(f"Kernel report: {kernel_path}")

    # 4. JSON summary
    json_path = generate_json_summary(output_dir, scan_data)
    log(f"JSON summary: {json_path}")

    log("Analysis complete!")

    return {
        "output_dir": str(output_dir),
        "excel": str(excel_path) if excel_path else None,
        "evidence": str(evidence_path),
        "kernel": str(kernel_path) if kernel_path else None,
        "json": str(json_path)
    }


def main():
    parser = argparse.ArgumentParser(description="Audit Report Analyzer")
    parser.add_argument("--archive", "-a", required=True, help="Path to audit report archive")
    parser.add_argument("--output", "-o", help="Output base directory")
    args = parser.parse_args()

    result = analyze_report(args.archive, args.output)

    print("\n" + "="*60)
    print("ANALYSIS COMPLETE")
    print("="*60)
    print(f"Output directory: {result['output_dir']}")
    print(f"  - Evidence report: {result['evidence']}")
    if result['excel']:
        print(f"  - Excel report: {result['excel']}")
    if result['kernel']:
        print(f"  - Kernel report: {result['kernel']}")
    print(f"  - JSON summary: {result['json']}")


if __name__ == "__main__":
    main()